#!/usr/bin/env python3
"""
Populate the SAVE Excel model's Ruler Table with data from research output .md files.
"""

import os
import re
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def extract_json_from_markdown(content):
    """Extract JSON block from markdown content."""
    match = re.search(r'```json\s*([\s\S]*?)\s*```', content)
    if match:
        try:
            return json.loads(match.group(1))
        except json.JSONDecodeError as e:
            print(f"  JSON parse error: {e}")
            return None
    return None

def populate_ruler_table(excel_path, research_dir):
    """Populate the Ruler Table tab with data from research .md files."""

    # Load the workbook
    wb = load_workbook(excel_path)
    ws = wb['Ruler Table']

    # Get all .md files
    md_files = sorted([f for f in os.listdir(research_dir) if f.endswith('.md')])

    print(f"Found {len(md_files)} research files")

    # Define styling
    input_fill_income = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    input_fill_asset = PatternFill(start_color="14532D", end_color="14532D", fill_type="solid")
    input_fill_default = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='4B5563'),
        right=Side(style='thin', color='4B5563'),
        top=Side(style='thin', color='4B5563'),
        bottom=Side(style='thin', color='4B5563')
    )

    # Default gold price
    default_gold_price = 2900

    # Process each file
    rulers_data = []
    for filename in md_files:
        filepath = os.path.join(research_dir, filename)
        print(f"Processing: {filename}")

        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()

        data = extract_json_from_markdown(content)
        if data:
            rulers_data.append(data)
            print(f"  Extracted: {data.get('ruler_name', 'Unknown')}")
        else:
            print(f"  WARNING: No JSON found in {filename}")

    # Sort by total wealth (calculated) descending
    def calc_total_wealth(d):
        try:
            agr = d['population'] * d['base_output'] * d['velocity']
            aii = agr * d['extraction_margin']
            income_wealth = aii * d['capitalization_multiple']
            infra = (d['population'] * d['base_output']) * 0.05
            asset_wealth = infra + d.get('liquid_treasury', 0) + d.get('imperial_real_estate', 0) + d.get('extraordinary_gains', 0)
            return income_wealth + asset_wealth
        except:
            return 0

    rulers_data.sort(key=calc_total_wealth, reverse=True)

    # Populate the Excel sheet starting at row 4
    for i, data in enumerate(rulers_data):
        row = i + 4  # Data starts at row 4

        print(f"Writing row {row}: {data.get('ruler_name', 'Unknown')}")

        # Column B: Ruler Name
        ws[f'B{row}'] = data.get('ruler_name', '')
        ws[f'B{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'B{row}'].fill = input_fill_default
        ws[f'B{row}'].border = thin_border

        # Column C: Year
        ws[f'C{row}'] = data.get('year', '')
        ws[f'C{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'C{row}'].fill = input_fill_default
        ws[f'C{row}'].border = thin_border

        # Column D: Population
        ws[f'D{row}'] = data.get('population', 0)
        ws[f'D{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'D{row}'].fill = input_fill_income
        ws[f'D{row}'].border = thin_border
        ws[f'D{row}'].number_format = '#,##0'

        # Column E: Base Output
        ws[f'E{row}'] = data.get('base_output', 1.0)
        ws[f'E{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'E{row}'].fill = input_fill_income
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].number_format = '0.00'

        # Column F: Velocity
        ws[f'F{row}'] = data.get('velocity', 4.0)
        ws[f'F{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'F{row}'].fill = input_fill_income
        ws[f'F{row}'].border = thin_border
        ws[f'F{row}'].number_format = '0.0'

        # Column G: Extraction Margin
        ws[f'G{row}'] = data.get('extraction_margin', 0.10)
        ws[f'G{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'G{row}'].fill = input_fill_income
        ws[f'G{row}'].border = thin_border
        ws[f'G{row}'].number_format = '0.00'

        # Column H: Capitalization Multiple
        ws[f'H{row}'] = data.get('capitalization_multiple', 20)
        ws[f'H{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'H{row}'].fill = input_fill_income
        ws[f'H{row}'].border = thin_border
        ws[f'H{row}'].number_format = '0'

        # Column I: Liquid Treasury
        ws[f'I{row}'] = data.get('liquid_treasury', 0)
        ws[f'I{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'I{row}'].fill = input_fill_asset
        ws[f'I{row}'].border = thin_border
        ws[f'I{row}'].number_format = '#,##0'

        # Column J: Imperial Real Estate
        ws[f'J{row}'] = data.get('imperial_real_estate', 0)
        ws[f'J{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'J{row}'].fill = input_fill_asset
        ws[f'J{row}'].border = thin_border
        ws[f'J{row}'].number_format = '#,##0'

        # Column K: Extraordinary Gains
        ws[f'K{row}'] = data.get('extraordinary_gains', 0)
        ws[f'K{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'K{row}'].fill = input_fill_asset
        ws[f'K{row}'].border = thin_border
        ws[f'K{row}'].number_format = '#,##0'

        # Column L: Gold Price
        ws[f'L{row}'] = default_gold_price
        ws[f'L{row}'].font = Font(size=10, color="FFFFFF")
        ws[f'L{row}'].fill = input_fill_default
        ws[f'L{row}'].border = thin_border
        ws[f'L{row}'].number_format = '#,##0'

    # Save the workbook
    wb.save(excel_path)
    print(f"\nExcel model updated: {excel_path}")
    print(f"Populated {len(rulers_data)} rulers in the Ruler Table tab")

if __name__ == "__main__":
    excel_path = "/home/markly2/claude_code/wealth_model/SAVE_Model.xlsx"
    research_dir = "/home/markly2/claude_code/wealth_model/research_outputs"

    populate_ruler_table(excel_path, research_dir)
