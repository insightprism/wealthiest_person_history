#!/usr/bin/env python3
"""
Create an Excel model for the Sovereign Asset Valuation Engine (SAVE)
This replicates the web calculator functionality in Excel with formulas.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

def create_save_model(filename="SAVE_Model.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "SAVE Calculator"

    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    section_font = Font(bold=True, size=12, color="F59E0B")  # Amber
    label_font = Font(size=11, color="D1D5DB")  # Gray
    input_font = Font(size=11, color="FFFFFF")
    output_font = Font(size=11, color="FFFFFF")
    result_font = Font(bold=True, size=14, color="F59E0B")

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # Dark gray
    input_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")  # Medium gray
    output_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    highlight_fill = PatternFill(start_color="78350F", end_color="78350F", fill_type="solid")  # Amber dark
    income_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")  # Blue
    asset_fill = PatternFill(start_color="14532D", end_color="14532D", fill_type="solid")  # Green

    thin_border = Border(
        left=Side(style='thin', color='4B5563'),
        right=Side(style='thin', color='4B5563'),
        top=Side(style='thin', color='4B5563'),
        bottom=Side(style='thin', color='4B5563')
    )

    # Set column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 3

    # =========================================
    # TITLE
    # =========================================
    ws.merge_cells('B2:F2')
    ws['B2'] = "Sovereign Asset Valuation Engine (SAVE)"
    ws['B2'].font = Font(bold=True, size=18, color="F59E0B")
    ws['B2'].fill = header_fill

    ws.merge_cells('B3:F3')
    ws['B3'] = "Calculate Real Wealth in Gold Ounces using the Consolidated Imperial Valuation (CIV) methodology"
    ws['B3'].font = Font(size=10, color="9CA3AF")
    ws['B3'].fill = header_fill

    # =========================================
    # INPUT SECTION (Left Column)
    # =========================================
    row = 5

    # Section Header
    ws[f'B{row}'] = "INPUT PARAMETERS"
    ws[f'B{row}'].font = section_font
    ws[f'B{row}'].fill = header_fill
    ws.merge_cells(f'B{row}:C{row}')
    row += 2

    # Ruler Name
    ws[f'B{row}'] = "Ruler/Empire Name"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = "Akbar the Great (Mughal Empire)"
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    ruler_name_cell = f'C{row}'
    row += 1

    # Year
    ws[f'B{row}'] = "Year"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = 1600
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    year_cell = f'C{row}'
    row += 2

    # Income Parameters Header
    ws[f'B{row}'] = "Income Parameters"
    ws[f'B{row}'].font = Font(bold=True, size=11, color="60A5FA")
    row += 1

    # Population
    ws[f'B{row}'] = "Population (N)"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = 110000000
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].number_format = '#,##0'
    pop_cell = f'C{row}'
    row += 1

    # Base Output
    ws[f'B{row}'] = "Base Output (B) oz/person/year"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = 0.75
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].number_format = '0.00'
    base_output_cell = f'C{row}'
    row += 1

    # Velocity
    ws[f'B{row}'] = "Velocity (V) [2.0-8.0]"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = 4.5
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].number_format = '0.0'
    velocity_cell = f'C{row}'
    row += 1

    # Extraction Margin
    ws[f'B{row}'] = "Extraction Margin (M) [0.01-0.50]"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = 0.10
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].number_format = '0.00'
    margin_cell = f'C{row}'
    row += 1

    # Capitalization Multiple
    ws[f'B{row}'] = "Capitalization Multiple (K) [1-100]"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = 20
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].number_format = '0'
    cap_mult_cell = f'C{row}'
    row += 2

    # Asset Parameters Header
    ws[f'B{row}'] = "Asset Parameters (Balance Sheet)"
    ws[f'B{row}'].font = Font(bold=True, size=11, color="34D399")
    row += 1

    # Liquid Treasury
    ws[f'B{row}'] = "Liquid Treasury (L) oz"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = 50000000
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].number_format = '#,##0'
    treasury_cell = f'C{row}'
    row += 1

    # Imperial Real Estate
    ws[f'B{row}'] = "Imperial Real Estate (R) oz"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = 10000000
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].number_format = '#,##0'
    real_estate_cell = f'C{row}'
    row += 1

    # Extraordinary Gains
    ws[f'B{row}'] = "Extraordinary Gains (E) oz"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = 5000000
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].number_format = '#,##0'
    extra_gains_cell = f'C{row}'
    row += 2

    # Valuation Settings Header
    ws[f'B{row}'] = "Valuation Settings"
    ws[f'B{row}'].font = Font(bold=True, size=11, color="F59E0B")
    row += 1

    # Gold Price
    ws[f'B{row}'] = "Gold Price ($/oz)"
    ws[f'B{row}'].font = label_font
    ws[f'C{row}'] = 2900
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].number_format = '#,##0'
    gold_price_cell = f'C{row}'

    # =========================================
    # OUTPUT SECTION (Right Column)
    # =========================================
    row = 5

    # Output Header with ruler name
    ws[f'E{row}'] = f'={ruler_name_cell}&" ("&{year_cell}&")"'
    ws[f'E{row}'].font = Font(bold=True, size=14, color="F59E0B")
    ws.merge_cells(f'E{row}:F{row}')
    row += 2

    # Income Valuation Section
    ws[f'E{row}'] = "Income Valuation (Operating Business)"
    ws[f'E{row}'].font = Font(bold=True, size=12, color="60A5FA")
    ws.merge_cells(f'E{row}:F{row}')
    row += 1

    # AGR
    ws[f'E{row}'] = "Annual Gross Revenue (AGR)"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'={pop_cell}*{base_output_cell}*{velocity_cell}'
    ws[f'F{row}'].font = output_font
    ws[f'F{row}'].number_format = '#,##0" oz"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    agr_cell = f'F{row}'
    row += 1

    # AII
    ws[f'E{row}'] = "Annual Imperial Income (AII)"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'={agr_cell}*{margin_cell}'
    ws[f'F{row}'].font = output_font
    ws[f'F{row}'].number_format = '#,##0" oz"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    aii_cell = f'F{row}'
    row += 1

    # Income Wealth
    ws[f'E{row}'] = "Income Wealth (AII × K)"
    ws[f'E{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'F{row}'] = f'={aii_cell}*{cap_mult_cell}'
    ws[f'F{row}'].font = Font(bold=True, size=11, color="F59E0B")
    ws[f'F{row}'].number_format = '#,##0" oz"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    income_wealth_cell = f'F{row}'
    row += 2

    # Asset Valuation Section
    ws[f'E{row}'] = "Asset Valuation (Balance Sheet)"
    ws[f'E{row}'].font = Font(bold=True, size=12, color="34D399")
    ws.merge_cells(f'E{row}:F{row}')
    row += 1

    # Infrastructure Floor
    ws[f'E{row}'] = "Infrastructure Floor (I)"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'=({pop_cell}*{base_output_cell})*0.05'
    ws[f'F{row}'].font = output_font
    ws[f'F{row}'].number_format = '#,##0" oz"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    infra_cell = f'F{row}'
    row += 1

    # Liquid Treasury (display)
    ws[f'E{row}'] = "Liquid Treasury (L)"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'={treasury_cell}'
    ws[f'F{row}'].font = output_font
    ws[f'F{row}'].number_format = '#,##0" oz"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 1

    # Imperial Real Estate (display)
    ws[f'E{row}'] = "Imperial Real Estate (R)"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'={real_estate_cell}'
    ws[f'F{row}'].font = output_font
    ws[f'F{row}'].number_format = '#,##0" oz"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 1

    # Extraordinary Gains (display)
    ws[f'E{row}'] = "Extraordinary Gains (E)"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'={extra_gains_cell}'
    ws[f'F{row}'].font = output_font
    ws[f'F{row}'].number_format = '#,##0" oz"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 1

    # Asset Wealth
    ws[f'E{row}'] = "Asset Wealth (I+L+R+E)"
    ws[f'E{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'F{row}'] = f'={infra_cell}+{treasury_cell}+{real_estate_cell}+{extra_gains_cell}'
    ws[f'F{row}'].font = Font(bold=True, size=11, color="34D399")
    ws[f'F{row}'].number_format = '#,##0" oz"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    asset_wealth_cell = f'F{row}'
    row += 2

    # Total Real Wealth Box
    ws[f'E{row}'] = "Total Real Wealth"
    ws[f'E{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'E{row}'].fill = highlight_fill
    ws[f'F{row}'] = f'={income_wealth_cell}+{asset_wealth_cell}'
    ws[f'F{row}'].font = Font(bold=True, size=14, color="F59E0B")
    ws[f'F{row}'].fill = highlight_fill
    ws[f'F{row}'].number_format = '#,##0" oz"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    total_wealth_cell = f'F{row}'
    row += 1

    # Total in full
    ws[f'E{row}'] = "(full number)"
    ws[f'E{row}'].font = Font(size=9, color="9CA3AF")
    ws[f'E{row}'].fill = highlight_fill
    ws[f'F{row}'] = f'={total_wealth_cell}'
    ws[f'F{row}'].font = Font(size=9, color="9CA3AF")
    ws[f'F{row}'].fill = highlight_fill
    ws[f'F{row}'].number_format = '#,##0" gold ounces"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 2

    # Nominal Dollar Value Box
    ws[f'E{row}'] = "Nominal Dollar Value"
    ws[f'E{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'E{row}'].fill = highlight_fill
    ws[f'F{row}'] = f'={total_wealth_cell}*{gold_price_cell}'
    ws[f'F{row}'].font = Font(bold=True, size=14, color="34D399")
    ws[f'F{row}'].fill = highlight_fill
    ws[f'F{row}'].number_format = '$#,##0,," T"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    nominal_value_cell = f'F{row}'
    row += 1

    # Nominal in full
    ws[f'E{row}'] = f'@ ${gold_price_cell}/oz'
    ws[f'E{row}'] = f'=CONCATENATE("@ $",TEXT({gold_price_cell},"#,##0"),"/oz")'
    ws[f'E{row}'].font = Font(size=9, color="9CA3AF")
    ws[f'E{row}'].fill = highlight_fill
    ws[f'F{row}'] = f'={nominal_value_cell}'
    ws[f'F{row}'].font = Font(size=9, color="9CA3AF")
    ws[f'F{row}'].fill = highlight_fill
    ws[f'F{row}'].number_format = '$#,##0" USD"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 2

    # =========================================
    # CFO DASHBOARD
    # =========================================
    ws[f'E{row}'] = "CFO Dashboard"
    ws[f'E{row}'].font = Font(bold=True, size=12, color="F59E0B")
    ws.merge_cells(f'E{row}:F{row}')
    row += 1

    # Income-to-Asset Ratio
    ws[f'E{row}'] = "Income-to-Asset Ratio"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'=IF({asset_wealth_cell}>0,{income_wealth_cell}/{asset_wealth_cell},0)'
    ws[f'F{row}'].font = Font(bold=True, size=11, color="60A5FA")
    ws[f'F{row}'].number_format = '0.00" x"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 1

    # Extraction Efficiency
    ws[f'E{row}'] = "Extraction Efficiency"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'={margin_cell}'
    ws[f'F{row}'].font = Font(bold=True, size=11, color="60A5FA")
    ws[f'F{row}'].number_format = '0.0%'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 1

    # Per Capita Wealth
    ws[f'E{row}'] = "Per Capita Wealth"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'=IF({pop_cell}>0,{total_wealth_cell}/{pop_cell},0)'
    ws[f'F{row}'].font = Font(bold=True, size=11, color="60A5FA")
    ws[f'F{row}'].number_format = '0.00" oz"'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 1

    # Annual Yield
    ws[f'E{row}'] = "Annual Yield (AII / Total Wealth)"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'=IF({total_wealth_cell}>0,{aii_cell}/{total_wealth_cell},0)'
    ws[f'F{row}'].font = Font(bold=True, size=11, color="60A5FA")
    ws[f'F{row}'].number_format = '0.00%'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 2

    # Wealth Composition
    ws[f'E{row}'] = "Wealth Composition"
    ws[f'E{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    row += 1

    ws[f'E{row}'] = "Income Wealth %"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'=IF({total_wealth_cell}>0,{income_wealth_cell}/{total_wealth_cell},0)'
    ws[f'F{row}'].font = output_font
    ws[f'F{row}'].number_format = '0.0%'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    income_pct_cell = f'F{row}'
    row += 1

    ws[f'E{row}'] = "Asset Wealth %"
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'] = f'=IF({total_wealth_cell}>0,{asset_wealth_cell}/{total_wealth_cell},0)'
    ws[f'F{row}'].font = output_font
    ws[f'F{row}'].number_format = '0.0%'
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 2

    # =========================================
    # FORMULAS REFERENCE
    # =========================================
    ws[f'E{row}'] = "Formulas Reference"
    ws[f'E{row}'].font = Font(bold=True, size=11, color="9CA3AF")
    row += 1

    formulas = [
        ("AGR", "N × B × V"),
        ("AII", "AGR × M"),
        ("Income Wealth", "AII × K"),
        ("Infrastructure (I)", "(N × B) × 0.05"),
        ("Asset Wealth", "I + L + R + E"),
        ("Total Wealth", "Income Wealth + Asset Wealth"),
    ]

    for name, formula in formulas:
        ws[f'E{row}'] = name
        ws[f'E{row}'].font = Font(size=9, color="6B7280")
        ws[f'F{row}'] = formula
        ws[f'F{row}'].font = Font(size=9, color="6B7280")
        row += 1

    # =========================================
    # Apply background to all cells
    # =========================================
    for r in range(1, row + 5):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            if cell.fill.start_color.index == '00000000' or cell.fill.start_color.index is None:
                cell.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")

    # =========================================
    # CREATE COMPARISON TABLE TAB
    # =========================================
    ws2 = wb.create_sheet(title="Comparison Table")

    # Define number of ruler columns (can add more by copying formulas)
    num_rulers = 20

    # Styles for the comparison sheet
    header_fill_dark = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    input_section_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    income_section_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    asset_section_fill = PatternFill(start_color="14532D", end_color="14532D", fill_type="solid")
    total_section_fill = PatternFill(start_color="78350F", end_color="78350F", fill_type="solid")
    cfo_section_fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")

    # Set column widths
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 35
    for i in range(num_rulers):
        col_letter = get_column_letter(i + 3)  # Start from column C
        ws2.column_dimensions[col_letter].width = 18

    # Title
    ws2.merge_cells('B2:L2')
    ws2['B2'] = "SAVE Comparison Table - Enter rulers in columns, results auto-calculate"
    ws2['B2'].font = Font(bold=True, size=14, color="F59E0B")
    ws2['B2'].fill = header_fill_dark

    # Build the row structure
    row = 4

    # Helper to add a row with label and formulas
    def add_row(label, row_type, formula_template=None, number_format='#,##0', is_input=False, section_fill=None):
        nonlocal row

        # Row number label
        ws2[f'A{row}'] = row - 3
        ws2[f'A{row}'].font = Font(size=9, color="6B7280")
        ws2[f'A{row}'].fill = header_fill_dark

        # Label
        ws2[f'B{row}'] = label
        if row_type == 'header':
            ws2[f'B{row}'].font = Font(bold=True, size=11, color="F59E0B")
        elif row_type == 'section':
            ws2[f'B{row}'].font = Font(bold=True, size=10, color="FFFFFF")
        elif row_type == 'total':
            ws2[f'B{row}'].font = Font(bold=True, size=11, color="F59E0B")
        else:
            ws2[f'B{row}'].font = Font(size=10, color="D1D5DB")

        if section_fill:
            ws2[f'B{row}'].fill = section_fill
        else:
            ws2[f'B{row}'].fill = header_fill_dark

        # Data columns
        for i in range(num_rulers):
            col = get_column_letter(i + 3)
            cell = ws2[f'{col}{row}']

            if row_type == 'header' or row_type == 'section':
                cell.fill = section_fill if section_fill else header_fill_dark
            elif is_input:
                cell.fill = input_section_fill
                cell.border = thin_border
            elif formula_template:
                # Apply formula
                cell.value = formula_template.replace('COL', col)
                if section_fill:
                    cell.fill = section_fill

            cell.font = Font(size=10, color="FFFFFF")
            if row_type == 'total':
                cell.font = Font(bold=True, size=11, color="F59E0B")
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal='right')

        current_row = row
        row += 1
        return current_row

    # =========================================
    # INPUT SECTION
    # =========================================
    add_row("INPUT PARAMETERS", 'header', section_fill=header_fill_dark)

    ruler_row = add_row("Ruler/Empire Name", 'input', is_input=True, number_format='@')
    year_row = add_row("Year", 'input', is_input=True, number_format='0')

    row += 1  # spacer
    add_row("Income Parameters", 'section', section_fill=income_section_fill)
    pop_row = add_row("Population (N)", 'input', is_input=True, number_format='#,##0')
    base_row = add_row("Base Output (B) oz/person/yr", 'input', is_input=True, number_format='0.00')
    vel_row = add_row("Velocity (V)", 'input', is_input=True, number_format='0.0')
    margin_row = add_row("Extraction Margin (M)", 'input', is_input=True, number_format='0.00')
    cap_row = add_row("Capitalization Multiple (K)", 'input', is_input=True, number_format='0')

    row += 1  # spacer
    add_row("Asset Parameters", 'section', section_fill=asset_section_fill)
    treasury_row = add_row("Liquid Treasury (L) oz", 'input', is_input=True, number_format='#,##0')
    realestate_row = add_row("Imperial Real Estate (R) oz", 'input', is_input=True, number_format='#,##0')
    extragains_row = add_row("Extraordinary Gains (E) oz", 'input', is_input=True, number_format='#,##0')

    row += 1  # spacer
    add_row("Settings", 'section', section_fill=header_fill_dark)
    goldprice_row = add_row("Gold Price ($/oz)", 'input', is_input=True, number_format='#,##0')

    # =========================================
    # OUTPUT SECTION - INCOME VALUATION
    # =========================================
    row += 1  # spacer
    add_row("INCOME VALUATION", 'header', section_fill=income_section_fill)

    agr_row = add_row("Annual Gross Revenue (AGR)", 'calc',
                      f'=IF(COL{pop_row}>0,COL{pop_row}*COL{base_row}*COL{vel_row},"")',
                      number_format='#,##0', section_fill=income_section_fill)

    aii_row = add_row("Annual Imperial Income (AII)", 'calc',
                      f'=IF(COL{agr_row}<>"",COL{agr_row}*COL{margin_row},"")',
                      number_format='#,##0', section_fill=income_section_fill)

    income_wealth_row = add_row("Income Wealth (AII × K)", 'total',
                                f'=IF(COL{aii_row}<>"",COL{aii_row}*COL{cap_row},"")',
                                number_format='#,##0', section_fill=income_section_fill)

    # =========================================
    # OUTPUT SECTION - ASSET VALUATION
    # =========================================
    row += 1  # spacer
    add_row("ASSET VALUATION", 'header', section_fill=asset_section_fill)

    infra_row = add_row("Infrastructure Floor (I)", 'calc',
                        f'=IF(COL{pop_row}>0,(COL{pop_row}*COL{base_row})*0.05,"")',
                        number_format='#,##0', section_fill=asset_section_fill)

    # Display asset inputs again for reference
    add_row("+ Liquid Treasury (L)", 'calc',
            f'=IF(COL{treasury_row}<>"",COL{treasury_row},"")',
            number_format='#,##0', section_fill=asset_section_fill)

    add_row("+ Imperial Real Estate (R)", 'calc',
            f'=IF(COL{realestate_row}<>"",COL{realestate_row},"")',
            number_format='#,##0', section_fill=asset_section_fill)

    add_row("+ Extraordinary Gains (E)", 'calc',
            f'=IF(COL{extragains_row}<>"",COL{extragains_row},"")',
            number_format='#,##0', section_fill=asset_section_fill)

    asset_wealth_row = add_row("Asset Wealth (I+L+R+E)", 'total',
                               f'=IF(COL{infra_row}<>"",COL{infra_row}+COL{treasury_row}+COL{realestate_row}+COL{extragains_row},"")',
                               number_format='#,##0', section_fill=asset_section_fill)

    # =========================================
    # TOTAL WEALTH
    # =========================================
    row += 1  # spacer
    add_row("TOTAL WEALTH", 'header', section_fill=total_section_fill)

    total_wealth_row = add_row("Total Real Wealth (oz)", 'total',
                               f'=IF(COL{income_wealth_row}<>"",COL{income_wealth_row}+COL{asset_wealth_row},"")',
                               number_format='#,##0', section_fill=total_section_fill)

    add_row("Total Wealth (Billions oz)", 'total',
            f'=IF(COL{total_wealth_row}<>"",COL{total_wealth_row}/1000000000,"")',
            number_format='0.00" B"', section_fill=total_section_fill)

    nominal_row = add_row("Nominal Value (USD)", 'total',
                          f'=IF(AND(COL{total_wealth_row}<>"",COL{goldprice_row}>0),COL{total_wealth_row}*COL{goldprice_row},"")',
                          number_format='$#,##0', section_fill=total_section_fill)

    add_row("Nominal Value (Trillions)", 'total',
            f'=IF(COL{nominal_row}<>"",COL{nominal_row}/1000000000000,"")',
            number_format='$0.00" T"', section_fill=total_section_fill)

    # =========================================
    # CFO METRICS
    # =========================================
    row += 1  # spacer
    add_row("CFO METRICS", 'header', section_fill=cfo_section_fill)

    add_row("Income-to-Asset Ratio", 'calc',
            f'=IF(AND(COL{income_wealth_row}<>"",COL{asset_wealth_row}>0),COL{income_wealth_row}/COL{asset_wealth_row},"")',
            number_format='0.00" x"', section_fill=cfo_section_fill)

    add_row("Extraction Efficiency", 'calc',
            f'=IF(COL{margin_row}<>"",COL{margin_row},"")',
            number_format='0.0%', section_fill=cfo_section_fill)

    add_row("Per Capita Wealth (oz)", 'calc',
            f'=IF(AND(COL{total_wealth_row}<>"",COL{pop_row}>0),COL{total_wealth_row}/COL{pop_row},"")',
            number_format='0.00', section_fill=cfo_section_fill)

    add_row("Annual Yield", 'calc',
            f'=IF(AND(COL{aii_row}<>"",COL{total_wealth_row}>0),COL{aii_row}/COL{total_wealth_row},"")',
            number_format='0.00%', section_fill=cfo_section_fill)

    add_row("Income Wealth %", 'calc',
            f'=IF(AND(COL{income_wealth_row}<>"",COL{total_wealth_row}>0),COL{income_wealth_row}/COL{total_wealth_row},"")',
            number_format='0.0%', section_fill=cfo_section_fill)

    add_row("Asset Wealth %", 'calc',
            f'=IF(AND(COL{asset_wealth_row}<>"",COL{total_wealth_row}>0),COL{asset_wealth_row}/COL{total_wealth_row},"")',
            number_format='0.0%', section_fill=cfo_section_fill)

    # =========================================
    # WEALTH COMPOSITION BREAKDOWN
    # =========================================
    row += 1  # spacer
    add_row("WEALTH COMPOSITION %", 'header', section_fill=header_fill_dark)

    add_row("Income Wealth %", 'calc',
            f'=IF(COL{total_wealth_row}>0,COL{income_wealth_row}/COL{total_wealth_row},"")',
            number_format='0.0%')

    add_row("Infrastructure %", 'calc',
            f'=IF(COL{total_wealth_row}>0,COL{infra_row}/COL{total_wealth_row},"")',
            number_format='0.0%')

    add_row("Liquid Treasury %", 'calc',
            f'=IF(COL{total_wealth_row}>0,COL{treasury_row}/COL{total_wealth_row},"")',
            number_format='0.0%')

    add_row("Real Estate %", 'calc',
            f'=IF(COL{total_wealth_row}>0,COL{realestate_row}/COL{total_wealth_row},"")',
            number_format='0.0%')

    add_row("Extraordinary Gains %", 'calc',
            f'=IF(COL{total_wealth_row}>0,COL{extragains_row}/COL{total_wealth_row},"")',
            number_format='0.0%')

    # Apply dark background to all cells
    for r in range(1, row + 3):
        for c in range(1, num_rulers + 4):
            cell = ws2.cell(row=r, column=c)
            if cell.fill.start_color.index == '00000000' or cell.fill.start_color.index is None:
                cell.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")

    # Add column headers (Ruler 1, Ruler 2, etc.)
    for i in range(num_rulers):
        col = get_column_letter(i + 3)
        ws2[f'{col}3'] = f'Ruler {i + 1}'
        ws2[f'{col}3'].font = Font(bold=True, size=10, color="F59E0B")
        ws2[f'{col}3'].fill = header_fill_dark
        ws2[f'{col}3'].alignment = Alignment(horizontal='center')

    # Freeze panes - freeze column A, B and row 3
    ws2.freeze_panes = 'C4'

    # Add sample data for first ruler
    ws2['C5'] = "Akbar the Great (Mughal Empire)"
    ws2['C6'] = 1600
    ws2['C8'] = 110000000
    ws2['C9'] = 0.75
    ws2['C10'] = 4.5
    ws2['C11'] = 0.10
    ws2['C12'] = 20
    ws2['C14'] = 50000000
    ws2['C15'] = 10000000
    ws2['C16'] = 5000000
    ws2['C18'] = 2900

    # Add second sample ruler
    ws2['D5'] = "Mansa Musa I (Mali Empire)"
    ws2['D6'] = 1324
    ws2['D8'] = 20000000
    ws2['D9'] = 0.8
    ws2['D10'] = 4.0
    ws2['D11'] = 0.15
    ws2['D12'] = 20
    ws2['D14'] = 16000000
    ws2['D15'] = 0
    ws2['D16'] = 0
    ws2['D18'] = 2900

    # =========================================
    # CREATE RULER TABLE TAB (Rows = Rulers, Columns = Variables)
    # =========================================
    ws3 = wb.create_sheet(title="Ruler Table")

    num_ruler_rows = 50  # Number of ruler rows available

    # Column definitions: (column_letter, header, width, is_input, number_format, formula_or_none)
    # Formulas use ROW to reference the current row
    columns = [
        ('A', '#', 4, False, '0', None),
        ('B', 'Ruler/Empire Name', 30, True, '@', None),
        ('C', 'Year', 8, True, '0', None),
        # Input parameters
        ('D', 'Population (N)', 15, True, '#,##0', None),
        ('E', 'Base Output (B)', 12, True, '0.00', None),
        ('F', 'Velocity (V)', 10, True, '0.0', None),
        ('G', 'Margin (M)', 10, True, '0.00', None),
        ('H', 'Cap Mult (K)', 10, True, '0', None),
        ('I', 'Treasury (L)', 15, True, '#,##0', None),
        ('J', 'Real Estate (R)', 15, True, '#,##0', None),
        ('K', 'Extra Gains (E)', 15, True, '#,##0', None),
        ('L', 'Gold Price', 12, True, '#,##0', None),
        # Calculated - Income
        ('M', 'AGR', 15, False, '#,##0', '=IF($D{row}>0,$D{row}*$E{row}*$F{row},"")'),
        ('N', 'AII', 15, False, '#,##0', '=IF($M{row}<>"",$M{row}*$G{row},"")'),
        ('O', 'Income Wealth', 18, False, '#,##0', '=IF($N{row}<>"",$N{row}*$H{row},"")'),
        # Calculated - Assets
        ('P', 'Infrastructure', 15, False, '#,##0', '=IF($D{row}>0,($D{row}*$E{row})*0.05,"")'),
        ('Q', 'Asset Wealth', 15, False, '#,##0', '=IF($P{row}<>"",$P{row}+$I{row}+$J{row}+$K{row},"")'),
        # Totals
        ('R', 'Total Wealth (oz)', 18, False, '#,##0', '=IF($O{row}<>"",$O{row}+$Q{row},"")'),
        ('S', 'Total (B oz)', 12, False, '0.00" B"', '=IF($R{row}<>"",$R{row}/1000000000,"")'),
        ('T', 'Nominal USD', 18, False, '$#,##0', '=IF(AND($R{row}<>"",$L{row}>0),$R{row}*$L{row},"")'),
        ('U', 'Nominal ($T)', 12, False, '$0.00" T"', '=IF($T{row}<>"",$T{row}/1000000000000,"")'),
        # CFO Metrics
        ('V', 'Income/Asset', 12, False, '0.00" x"', '=IF(AND($O{row}<>"",$Q{row}>0),$O{row}/$Q{row},"")'),
        ('W', 'Per Capita', 12, False, '0.00', '=IF(AND($R{row}<>"",$D{row}>0),$R{row}/$D{row},"")'),
        ('X', 'Annual Yield', 10, False, '0.00%', '=IF(AND($N{row}<>"",$R{row}>0),$N{row}/$R{row},"")'),
        ('Y', 'Income %', 10, False, '0.0%', '=IF(AND($O{row}<>"",$R{row}>0),$O{row}/$R{row},"")'),
        ('Z', 'Asset %', 10, False, '0.0%', '=IF(AND($Q{row}<>"",$R{row}>0),$Q{row}/$R{row},"")'),
    ]

    # Set column widths and create headers
    for col_letter, header, width, is_input, num_fmt, formula in columns:
        ws3.column_dimensions[col_letter].width = width

    # Title row
    ws3.merge_cells('B1:U1')
    ws3['B1'] = "SAVE Ruler Table - Enter one ruler per row, results auto-calculate"
    ws3['B1'].font = Font(bold=True, size=14, color="F59E0B")
    ws3['B1'].fill = header_fill_dark

    # Section headers row 2
    ws3.merge_cells('B2:C2')
    ws3['B2'] = "Identity"
    ws3['B2'].font = Font(bold=True, size=10, color="FFFFFF")
    ws3['B2'].fill = header_fill_dark
    ws3['B2'].alignment = Alignment(horizontal='center')

    ws3.merge_cells('D2:H2')
    ws3['D2'] = "Income Parameters"
    ws3['D2'].font = Font(bold=True, size=10, color="60A5FA")
    ws3['D2'].fill = income_section_fill
    ws3['D2'].alignment = Alignment(horizontal='center')

    ws3.merge_cells('I2:K2')
    ws3['I2'] = "Asset Parameters"
    ws3['I2'].font = Font(bold=True, size=10, color="34D399")
    ws3['I2'].fill = asset_section_fill
    ws3['I2'].alignment = Alignment(horizontal='center')

    ws3['L2'] = "Settings"
    ws3['L2'].font = Font(bold=True, size=10, color="F59E0B")
    ws3['L2'].fill = header_fill_dark
    ws3['L2'].alignment = Alignment(horizontal='center')

    ws3.merge_cells('M2:O2')
    ws3['M2'] = "Income Valuation"
    ws3['M2'].font = Font(bold=True, size=10, color="60A5FA")
    ws3['M2'].fill = income_section_fill
    ws3['M2'].alignment = Alignment(horizontal='center')

    ws3.merge_cells('P2:Q2')
    ws3['P2'] = "Asset Valuation"
    ws3['P2'].font = Font(bold=True, size=10, color="34D399")
    ws3['P2'].fill = asset_section_fill
    ws3['P2'].alignment = Alignment(horizontal='center')

    ws3.merge_cells('R2:U2')
    ws3['R2'] = "Total Wealth"
    ws3['R2'].font = Font(bold=True, size=10, color="F59E0B")
    ws3['R2'].fill = total_section_fill
    ws3['R2'].alignment = Alignment(horizontal='center')

    ws3.merge_cells('V2:Z2')
    ws3['V2'] = "CFO Metrics"
    ws3['V2'].font = Font(bold=True, size=10, color="A78BFA")
    ws3['V2'].fill = cfo_section_fill
    ws3['V2'].alignment = Alignment(horizontal='center')

    # Column headers row 3
    for col_letter, header, width, is_input, num_fmt, formula in columns:
        cell = ws3[f'{col_letter}3']
        cell.value = header
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = header_fill_dark
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Add color coding to input vs output headers
    for col_letter, header, width, is_input, num_fmt, formula in columns:
        cell = ws3[f'{col_letter}3']
        if is_input and col_letter not in ['A']:
            if col_letter in ['D', 'E', 'F', 'G', 'H']:
                cell.fill = income_section_fill
            elif col_letter in ['I', 'J', 'K']:
                cell.fill = asset_section_fill
            elif col_letter == 'L':
                cell.fill = header_fill_dark
            else:
                cell.fill = input_section_fill
        elif col_letter in ['M', 'N', 'O']:
            cell.fill = income_section_fill
        elif col_letter in ['P', 'Q']:
            cell.fill = asset_section_fill
        elif col_letter in ['R', 'S', 'T', 'U']:
            cell.fill = total_section_fill
        elif col_letter in ['V', 'W', 'X', 'Y', 'Z']:
            cell.fill = cfo_section_fill

    # Data rows
    for i in range(num_ruler_rows):
        row_num = i + 4  # Data starts at row 4

        for col_letter, header, width, is_input, num_fmt, formula in columns:
            cell = ws3[f'{col_letter}{row_num}']

            if col_letter == 'A':
                cell.value = i + 1
                cell.font = Font(size=9, color="6B7280")
                cell.fill = header_fill_dark
            elif is_input:
                # Input cell styling
                if col_letter in ['D', 'E', 'F', 'G', 'H']:
                    cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                elif col_letter in ['I', 'J', 'K']:
                    cell.fill = PatternFill(start_color="14532D", end_color="14532D", fill_type="solid")
                else:
                    cell.fill = input_section_fill
                cell.border = thin_border
                cell.font = Font(size=10, color="FFFFFF")
            elif formula:
                # Output cell with formula
                cell.value = formula.format(row=row_num)
                if col_letter in ['M', 'N', 'O']:
                    cell.fill = PatternFill(start_color="172554", end_color="172554", fill_type="solid")
                elif col_letter in ['P', 'Q']:
                    cell.fill = PatternFill(start_color="0F2F1E", end_color="0F2F1E", fill_type="solid")
                elif col_letter in ['R', 'S', 'T', 'U']:
                    cell.fill = PatternFill(start_color="451A03", end_color="451A03", fill_type="solid")
                    cell.font = Font(bold=True, size=10, color="F59E0B")
                elif col_letter in ['V', 'W', 'X', 'Y', 'Z']:
                    cell.fill = PatternFill(start_color="2E1065", end_color="2E1065", fill_type="solid")
                    cell.font = Font(size=10, color="FFFFFF")
                else:
                    cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
                    cell.font = Font(size=10, color="FFFFFF")

                # Override font for total columns
                if col_letter not in ['R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
                    cell.font = Font(size=10, color="FFFFFF")

            cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal='right' if col_letter != 'B' else 'left')

    # Freeze panes - freeze row 3 and column A
    ws3.freeze_panes = 'B4'

    # Add sample data
    sample_data = [
        ("Akbar the Great (Mughal Empire)", 1600, 110000000, 0.75, 4.5, 0.10, 20, 50000000, 10000000, 5000000, 2900),
        ("Mansa Musa I (Mali Empire)", 1324, 20000000, 0.8, 4.0, 0.15, 20, 16000000, 0, 0, 2900),
        ("Emperor Shenzong (Song Dynasty)", 1080, 100000000, 1.2, 6.0, 0.12, 22, 80000000, 100000000, 0, 2900),
        ("Augustus Caesar (Roman Empire)", 14, 55000000, 0.8, 4.5, 0.15, 18, 30000000, 20000000, 50000000, 2900),
        ("Genghis Khan (Mongol Empire)", 1227, 60000000, 0.5, 3.0, 0.30, 15, 50000000, 10000000, 100000000, 2900),
    ]

    for i, data in enumerate(sample_data):
        row_num = i + 4
        ws3[f'B{row_num}'] = data[0]
        ws3[f'C{row_num}'] = data[1]
        ws3[f'D{row_num}'] = data[2]
        ws3[f'E{row_num}'] = data[3]
        ws3[f'F{row_num}'] = data[4]
        ws3[f'G{row_num}'] = data[5]
        ws3[f'H{row_num}'] = data[6]
        ws3[f'I{row_num}'] = data[7]
        ws3[f'J{row_num}'] = data[8]
        ws3[f'K{row_num}'] = data[9]
        ws3[f'L{row_num}'] = data[10]

    # Apply dark background to empty cells
    for r in range(1, num_ruler_rows + 5):
        for c in range(1, 27):  # A to Z
            cell = ws3.cell(row=r, column=c)
            if cell.fill.start_color.index == '00000000' or cell.fill.start_color.index is None:
                cell.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")

    # =========================================
    # CREATE READ ME TAB
    # =========================================
    ws4 = wb.create_sheet(title="Read Me", index=0)  # Insert at beginning

    # Styles
    title_font = Font(bold=True, size=18, color="F59E0B")
    h1_font = Font(bold=True, size=14, color="F59E0B")
    h2_font = Font(bold=True, size=12, color="60A5FA")
    h3_font = Font(bold=True, size=11, color="34D399")
    body_font = Font(size=10, color="D1D5DB")
    formula_font = Font(size=10, color="A78BFA", italic=True)
    highlight_font = Font(bold=True, size=10, color="F59E0B")
    dark_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")

    # Set column widths
    ws4.column_dimensions['A'].width = 3
    ws4.column_dimensions['B'].width = 25
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 60
    ws4.column_dimensions['E'].width = 30
    ws4.column_dimensions['F'].width = 3

    row = 2

    def add_text(text, font=body_font, col='B', merge_to='E'):
        nonlocal row
        ws4[f'{col}{row}'] = text
        ws4[f'{col}{row}'].font = font
        ws4[f'{col}{row}'].fill = dark_fill
        ws4[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        if merge_to:
            ws4.merge_cells(f'{col}{row}:{merge_to}{row}')
        row += 1

    def add_blank():
        nonlocal row
        row += 1

    def add_variable(name, symbol, description, formula=""):
        nonlocal row
        ws4[f'B{row}'] = name
        ws4[f'B{row}'].font = highlight_font
        ws4[f'B{row}'].fill = dark_fill

        ws4[f'C{row}'] = symbol
        ws4[f'C{row}'].font = formula_font
        ws4[f'C{row}'].fill = dark_fill

        ws4[f'D{row}'] = description
        ws4[f'D{row}'].font = body_font
        ws4[f'D{row}'].fill = dark_fill
        ws4[f'D{row}'].alignment = Alignment(wrap_text=True, vertical='top')

        ws4[f'E{row}'] = formula
        ws4[f'E{row}'].font = formula_font
        ws4[f'E{row}'].fill = dark_fill

        row += 1

    # =========================================
    # CONTENT
    # =========================================

    # Title
    add_text("SOVEREIGN ASSET VALUATION ENGINE (SAVE)", title_font)
    add_text("User Guide & Model Documentation", Font(size=12, color="9CA3AF"))
    add_blank()

    # Overview
    add_text("1. OVERVIEW", h1_font)
    add_text("The Sovereign Asset Valuation Engine (SAVE) calculates the 'Real Wealth' of historical sovereign rulers using the Consolidated Imperial Valuation (CIV) methodology. All values are expressed in Gold Ounces as an inflation-neutral unit of account.")
    add_blank()
    add_text("The model combines two approaches:")
    add_text("  • Income Valuation - Values the ruler's ongoing revenue stream (like valuing a business)")
    add_text("  • Asset Valuation - Values the ruler's accumulated assets (like a balance sheet)")
    add_blank()

    # Tabs Description
    add_text("2. SPREADSHEET TABS", h1_font)
    add_blank()
    add_text("SAVE Calculator", h2_font)
    add_text("Single-ruler calculator with inputs on the left and results on the right. Good for detailed analysis of one ruler at a time.")
    add_blank()
    add_text("Comparison Table", h2_font)
    add_text("Variables in rows, rulers in columns. Good for comparing specific metrics across multiple rulers. Enter data in columns C, D, E, etc.")
    add_blank()
    add_text("Ruler Table", h2_font)
    add_text("Rulers in rows, variables in columns. Traditional spreadsheet format - one ruler per row. Best for building a database of many rulers. Pre-populated with 16 historical figures.")
    add_blank()

    # Core Formula
    add_text("3. CORE FORMULAS", h1_font)
    add_blank()
    add_text("Total Real Wealth = Income Wealth + Asset Wealth", formula_font)
    add_blank()
    add_text("Income Wealth Calculation:", h3_font)
    add_text("  AGR = N × B × V", formula_font)
    add_text("  AII = AGR × M", formula_font)
    add_text("  Income Wealth = AII × K", formula_font)
    add_blank()
    add_text("Asset Wealth Calculation:", h3_font)
    add_text("  Infrastructure (I) = (N × B) × 0.05", formula_font)
    add_text("  Asset Wealth = I + L + R + E", formula_font)
    add_blank()

    # Input Variables
    add_text("4. INPUT VARIABLES", h1_font)
    add_blank()
    add_text("Identity", h2_font)
    add_variable("Ruler/Empire Name", "-", "Name of the sovereign ruler and their empire/kingdom", "Text")
    add_variable("Year", "-", "Peak year of wealth measurement (use negative for BC)", "Number")
    add_blank()

    add_text("Income Parameters", h2_font)
    add_variable("Population", "N", "Total population under the ruler's control. For private individuals (non-sovereigns), use 1.", "Number (persons)")
    add_variable("Base Output", "B", "Economic output per person per year in gold ounce equivalents. Default 1.0 oz. Lower for pre-industrial/subsistence economies (0.3-0.8), higher for advanced economies (1.0-1.5).", "oz/person/year")
    add_variable("Velocity", "V", "Economic velocity multiplier reflecting how actively wealth circulates. Range 2.0-8.0. Lower for barter/subsistence economies, higher for monetized economies with banking.", "Multiplier (2.0-8.0)")
    add_variable("Extraction Margin", "M", "Percentage of economic output the ruler extracts through taxes, tribute, monopolies. Range 0.01-0.50 (1%-50%). Typical range 0.10-0.25.", "Decimal (0.01-0.50)")
    add_variable("Capitalization Multiple", "K", "Multiple applied to annual income to get total value (like P/E ratio). Default 20. Lower for unstable regimes (5-15), higher for stable empires (20-25).", "Multiplier (1-100)")
    add_blank()

    add_text("Asset Parameters", h2_font)
    add_variable("Liquid Treasury", "L", "Gold, silver, jewels, and cash reserves in the ruler's treasury. Expressed in gold ounce equivalents.", "Gold ounces")
    add_variable("Imperial Real Estate", "R", "Value of crown lands, palaces, mines, monopolies, and productive assets. Expressed in gold ounce equivalents.", "Gold ounces")
    add_variable("Extraordinary Gains", "E", "One-time windfalls from conquest, inheritance, or plunder. Not recurring income.", "Gold ounces")
    add_blank()

    add_text("Valuation Settings", h2_font)
    add_variable("Gold Price", "$/oz", "Current gold price in USD per troy ounce. Used to convert gold ounces to nominal USD value. Default $2,900.", "USD per oz")
    add_blank()

    # Output Variables
    add_text("5. OUTPUT VARIABLES (Calculated)", h1_font)
    add_blank()

    add_text("Income Valuation", h2_font)
    add_variable("Annual Gross Revenue", "AGR", "Total economic output under the ruler's control.", "N × B × V")
    add_variable("Annual Imperial Income", "AII", "Portion of AGR extracted by the ruler annually.", "AGR × M")
    add_variable("Income Wealth", "-", "Capitalized value of the income stream.", "AII × K")
    add_blank()

    add_text("Asset Valuation", h2_font)
    add_variable("Infrastructure Floor", "I", "Minimum infrastructure value (5% of base economic output).", "(N × B) × 0.05")
    add_variable("Asset Wealth", "-", "Sum of all asset components.", "I + L + R + E")
    add_blank()

    add_text("Total Wealth", h2_font)
    add_variable("Total Real Wealth", "-", "Combined income and asset wealth in gold ounces.", "Income Wealth + Asset Wealth")
    add_variable("Total (Billions)", "-", "Total wealth expressed in billions of gold ounces.", "Total / 1,000,000,000")
    add_variable("Nominal USD", "-", "Total wealth converted to US dollars.", "Total × Gold Price")
    add_variable("Nominal (Trillions)", "-", "Nominal value in trillions of USD.", "Nominal / 1,000,000,000,000")
    add_blank()

    # CFO Metrics
    add_text("6. CFO METRICS (Calculated)", h1_font)
    add_blank()
    add_variable("Income-to-Asset Ratio", "-", "Ratio of income wealth to asset wealth. Higher = more income-driven wealth.", "Income Wealth / Asset Wealth")
    add_variable("Extraction Efficiency", "-", "Same as Extraction Margin (M). Shows % of economy captured.", "M (as percentage)")
    add_variable("Per Capita Wealth", "-", "Total wealth divided by population. Wealth per person.", "Total Wealth / N")
    add_variable("Annual Yield", "-", "Annual income as % of total wealth. Like dividend yield.", "AII / Total Wealth")
    add_variable("Income Wealth %", "-", "Percentage of total wealth from income valuation.", "Income Wealth / Total")
    add_variable("Asset Wealth %", "-", "Percentage of total wealth from asset valuation.", "Asset Wealth / Total")
    add_blank()

    # Usage Guide
    add_text("7. HOW TO USE", h1_font)
    add_blank()
    add_text("For Single Ruler Analysis:", h3_font)
    add_text("  1. Go to 'SAVE Calculator' tab")
    add_text("  2. Enter ruler name and year")
    add_text("  3. Fill in Income Parameters (N, B, V, M, K)")
    add_text("  4. Fill in Asset Parameters (L, R, E)")
    add_text("  5. Adjust gold price if needed")
    add_text("  6. Results appear automatically on the right")
    add_blank()

    add_text("For Multiple Ruler Comparison:", h3_font)
    add_text("  1. Go to 'Ruler Table' tab")
    add_text("  2. Each row is one ruler - fill in columns B through L (inputs)")
    add_text("  3. Columns M through Z calculate automatically")
    add_text("  4. Sort by any column to rank rulers")
    add_text("  5. Use 'Comparison Table' tab for vertical comparison view")
    add_blank()

    # Special Cases
    add_text("8. SPECIAL CASES", h1_font)
    add_blank()
    add_text("Private Individuals (Non-Sovereigns):", h3_font)
    add_text("For private bankers, merchants, or families (e.g., Rothschilds, Fugger, Astor) who did not rule populations:")
    add_text("  • Set Population (N) = 1")
    add_text("  • Set Base Output (B) = 0.01")
    add_text("  • Set Extraction Margin (M) = 0.01")
    add_text("  • Set Capitalization Multiple (K) = 1")
    add_text("  • Enter their actual wealth in Liquid Treasury (L) and Real Estate (R)")
    add_text("  • Their wealth will be purely asset-based")
    add_blank()

    add_text("Ancient Rulers with Uncertain Data:", h3_font)
    add_text("  • Use conservative population estimates")
    add_text("  • Lower Base Output for pre-industrial economies (0.3-0.5)")
    add_text("  • Lower Velocity for barter economies (2.0-3.0)")
    add_text("  • Lower Capitalization Multiple for unstable regimes (5-15)")
    add_blank()

    # Interpretation
    add_text("9. INTERPRETING RESULTS", h1_font)
    add_blank()
    add_text("Wealth Composition:", h3_font)
    add_text("  • High Income % (>80%): Wealth based on economic control (large empires)")
    add_text("  • High Asset % (>50%): Wealth based on accumulated treasure (conquest, inheritance)")
    add_text("  • Balanced: Mix of ongoing revenue and stored wealth")
    add_blank()

    add_text("Income-to-Asset Ratio:", h3_font)
    add_text("  • >10x: Empire primarily valued for its economic output")
    add_text("  • 1-10x: Balanced between income and assets")
    add_text("  • <1x: Wealth primarily from accumulated assets (typical for private individuals)")
    add_blank()

    add_text("Per Capita Wealth:", h3_font)
    add_text("  • Higher = more extractive or more developed economy")
    add_text("  • Compare rulers with similar time periods for meaningful comparison")
    add_blank()

    # Caveats
    add_text("10. CAVEATS & LIMITATIONS", h1_font)
    add_blank()
    add_text("  • Historical data is often uncertain - treat results as estimates")
    add_text("  • Model designed for sovereign rulers - private individuals need modified approach")
    add_text("  • Gold ounce equivalents are approximations for non-monetary economies")
    add_text("  • Nominal USD values depend on gold price assumption")
    add_text("  • Cross-era comparisons have inherent limitations")
    add_text("  • Model assumes stable rule - may overvalue unstable empires")
    add_blank()

    # Version
    add_text("11. VERSION", h1_font)
    add_text("SAVE Model v1.0 - Consolidated Imperial Valuation (CIV) Methodology")
    add_text("Created: 2026")
    add_blank()

    # Apply dark fill to all cells
    for r in range(1, row + 10):
        for c in range(1, 7):
            cell = ws4.cell(row=r, column=c)
            if cell.fill.start_color.index == '00000000' or cell.fill.start_color.index is None:
                cell.fill = dark_fill

    # Save the workbook
    wb.save(filename)
    print(f"Excel model saved to: {filename}")
    return filename

if __name__ == "__main__":
    create_save_model("/home/markly2/claude_code/wealth_model/SAVE_Model.xlsx")
